# coding: utf-8
# !/usr/bin/python

# This file is part of the FlOpEDT/FlOpScheduler project.
# Copyright (c) 2017
# Authors: Iulian Ober, Paul Renaud-Goud, Pablo Seban, et al.
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU Affero General Public License as
# published by the Free Software Foundation, either version 3 of the
# License, or (at your option) any later version.
#
# This program is distributed in the hope that it will be useful, but
# WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE. See the GNU
# Affero General Public License for more details.
#
# You should have received a copy of the GNU Affero General Public
# License along with this program. If not, see
# <http://www.gnu.org/licenses/>.
#
# You can be released from the requirements of the license by purchasing
# a commercial license. Buying such a license is mandatory as soon as
# you develop activities involving the FlOpEDT/FlOpScheduler software
# without disclosing the source code of your own applications.

from openpyxl import load_workbook

from base.models import Group, Module, Period, CourseType

from copy import copy


def as_text(value):
    if value is None:
        return ""
    return str(value)


def column_letter(col):
    if col <= 26:
        return chr(64 + col)
    else:
        return chr(64 + (col - 1) // 26) + chr(65 + (col - 1) % 26)


def append_row(work_sheet, rows_to_append, row_number, rank, until):
    """
    Add a copy of 'rows_to_append[row_number]' to work_sheet with 'until' columns
    """
    row = rows_to_append[row_number - 1]
    for cell in row[:until]:
        new_cell = work_sheet.cell(row=rank, column=cell.col_idx, value=cell.value)
        if cell.has_style:
            new_cell.font = copy(cell.font)
            new_cell.border = copy(cell.border)
            new_cell.fill = copy(cell.fill)
            new_cell.number_format = copy(cell.number_format)
            new_cell.protection = copy(cell.protection)
            new_cell.alignment = copy(cell.alignment)


empty_bookname = 'misc/deploy_database/empty_planif_file_with_recap.xlsx'


def make_planif_file(department, empty_bookname=empty_bookname):
    new_book = load_workbook(filename=empty_bookname)
    empty_rows = list(new_book['empty'].rows)
    recap_rows = list(new_book['empty_recap'].rows)
    new_book.create_sheet('Recap')
    last_row = {}
    last_column_letter = {}
    # We go through each period and create a sheet for each period
    for p in Period.objects.filter(department=department):
        new_book.create_sheet(p.name)
        sheet = new_book[p.name]
        ################ Writing line 1 with weeks ################
        rank = 1
        FIRST_WEEK_COL = 8
        week_col = FIRST_WEEK_COL

        if p.starting_week < p.ending_week:
            weeks = p.ending_week - p.starting_week + 1
            cols = weeks + 7
            append_row(sheet, empty_rows, 1, rank, cols)
            for i in range(p.starting_week, p.ending_week + 1):
                sheet.cell(row=rank, column=week_col).value = i
                week_col += 1

        else:
            weeks = (53 - p.starting_week) + p.ending_week
            cols = weeks + 7
            append_row(sheet, empty_rows, 1, rank, cols)
            for i in range(p.starting_week, 53):
                sheet.cell(row=rank, column=week_col).value = i
                week_col += 1

            for i in range(1, p.ending_week + 1):
                sheet.cell(row=rank, column=week_col).value = i
                week_col += 1
        rank += 1
        append_row(sheet, empty_rows, 4, rank, cols)
        last_column_letter[p] = column_letter(cols)


        ################ A line per module per CourseType ################
        for mod in Module.objects.filter(period=p):
            for ct in CourseType.objects.filter(department=department):
                rank += 1
                append_row(sheet, empty_rows, 2, rank, cols)
                sheet.cell(row=rank, column=1).value = mod.abbrev
                sheet.cell(row=rank, column=2).value = '=$C%d&"_"&$E%d' % (rank, rank)
                sheet.cell(row=rank, column=3).value = ct.name
                sheet.cell(row=rank, column=4).value = ct.duration
                sheet.cell(row=rank, column=5).value = 'Prof'
                sheet.cell(row=rank, column=6).value = 'Type de Salle'
                sheet.cell(row=rank, column=7).value = 'Groupes'

                if Group.objects.filter(type__in=ct.group_types.all(), train_prog=mod.train_prog).count() > 0:
                    for g in Group.objects.filter(type__in=ct.group_types.all(), train_prog=mod.train_prog):
                        rank += 1
                        append_row(sheet, empty_rows, 3, rank, cols)
                        sheet.cell(row=rank, column=1).value = mod.abbrev
                        sheet.cell(row=rank, column=2).value = '=$C%d&"_"&$E%d' % (rank, rank)
                        sheet.cell(row=rank, column=3).value = ct.name
                        sheet.cell(row=rank, column=4).value = ct.duration
                        sheet.cell(row=rank, column=7).value = g.nom
                else:
                    rank += 1
                    append_row(sheet, empty_rows, 3, rank, cols)
                    sheet.cell(row=rank, column=1).value = mod.abbrev
                    sheet.cell(row=rank, column=2).value = '=$C%d&"_"&$E%d' % (rank, rank)
                    sheet.cell(row=rank, column=3).value = ct.name
                    sheet.cell(row=rank, column=4).value = ct.duration

            ################ Separating each course with a black line ################
            rank += 1
            append_row(sheet, empty_rows, 4, rank, cols)

        ############ TOTAL line ############
        ligne_finale = rank - 1
        rank += 1
        append_row(sheet, empty_rows, 5, rank, cols)
        for week_col in range(FIRST_WEEK_COL, cols+1):
            cl = column_letter(week_col)
            sheet.cell(row=rank, column=week_col).value = \
                '=SUMPRODUCT((D$3:D$%d)*(%s$3:%s$%d)*(G$3:G$%d="Groupes"))/60' \
                % (ligne_finale, cl, cl, ligne_finale, ligne_finale)
        rank += 1
        append_row(sheet, empty_rows, 4, rank, cols)

        ############ Other TOTAL lines ############
        rank += 1
        append_row(sheet, empty_rows, 7, rank, cols)
        for week_col in range(FIRST_WEEK_COL, cols + 1):
            cl = column_letter(week_col)
            sheet.cell(row=rank, column=week_col).value = '=%s1' % cl

        for ct in CourseType.objects.filter(department=department):
            rank += 1
            append_row(sheet, empty_rows, 8, rank, cols)
            sheet.cell(row=rank, column=2).value = '=$F%d&"_"&$E%d' % (rank, rank)
            sheet.cell(row=rank, column=5).value = "='Recap'!$B$1"
            sheet.cell(row=rank, column=6).value = ct.name
            sheet.cell(row=rank, column=7).value = '=SUM(H%d:%s%d)' % (rank, last_column_letter[p], rank)
            for week_col in range(FIRST_WEEK_COL, cols + 1):
                cl = column_letter(week_col)
                sheet.cell(row=rank, column=week_col).value =\
                    '=SUMPRODUCT((D$3:D$%d)*(%s$3:%s$%d)*($B$3:$B$%d=$B%d))/60' \
                    % (ligne_finale, cl, cl, ligne_finale, ligne_finale, rank)
        rank += 1
        append_row(sheet, empty_rows, 9, rank, cols)
        nb_ct = CourseType.objects.filter(department=department).count()
        for week_col in range(FIRST_WEEK_COL-1, cols + 1):
            cl = column_letter(week_col)
            sheet.cell(row=rank, column=week_col).value = \
                '=SUM(%s%d:%s%d)' % (cl, rank - nb_ct, cl, rank - 1)
        last_row[p.name]=rank

        ############ Adapting column widths ############
        for col in sheet.columns:
            length = len(as_text(col[0].value))
            adjusted_length = (length + 2) * 1.2
            sheet.column_dimensions[col[0].column].width = adjusted_length
        sheet.column_dimensions['B'].hidden = True


    ############ Make recap sheet ############
    sheet = new_book['Recap']
    rank = 1
    recap_col_nb = 48
    nb_per = Period.objects.filter(department=department).count()
    append_row(sheet, recap_rows, 1, rank, recap_col_nb)
    rank += 1
    for p in Period.objects.filter(department=department):
        append_row(sheet, recap_rows, 2, rank, recap_col_nb)
        sheet.cell(row=rank, column=1).value = p.name
        sheet.cell(row=rank, column=2).value = '=SUM($C%d:AV%d)' % (rank, rank)
        for week_col in range(3, recap_col_nb + 1):
            cl = column_letter(week_col)
            sheet.cell(row=rank, column=week_col).value = \
                '=SUMPRODUCT((%s!$H$%d:$%s$%d)*(%s!$H$1:$%s$1=%s$1))' % \
                (p.name, last_row[p.name], last_column_letter[p], last_row[p.name], p.name,  last_column_letter[p], cl)
            # '=SUMIF(%s!$H$1:$%s$1;%s$1;%s!$H$%d:$%s$%d)' (p.name, last_column_letter[p], cl, p.name, last_row[p.name], last_column_letter[p], last_row[p.name])

        rank += 1
    append_row(sheet, recap_rows, 3, rank, recap_col_nb)
    for week_col in range(2, recap_col_nb+1):
        cl = column_letter(week_col)
        sheet.cell(row=rank, column=week_col).value = \
            '=SUM(%s%d:%s%d)' % (cl, rank - nb_per, cl, rank - 1)
    rank += 1

    ############ Adapting column widths ############
    for col in sheet.columns:
        length = len(as_text(col[0].value))
        if length == 2:
            length = 1
        adjusted_length = (length + 2) * 1.2
        sheet.column_dimensions[col[0].column].width = adjusted_length
    new_book.remove(new_book['empty_recap'])

    new_book.remove(new_book['empty'])
    filename = 'misc/deploy_database/planif_file_' + department.abbrev + '.xlsx'
    new_book.save(filename=filename)
