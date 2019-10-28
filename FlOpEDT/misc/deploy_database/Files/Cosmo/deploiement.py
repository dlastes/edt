#!/usr/bin/env python3
# -*- coding: utf-8 -*-

# This file is part of the FlOpEDT/FlOpScheduler project.
# Copyright (c) 2017
# Authors: Iulian Ober, Paul Renaud-Goud, Pablo Seban, et al.
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU Affero General Public License as
# published by the Free Software Foundation, either version 3 of the
# License, or (at your option) any later version.
#
# This program is distributed in the hope that it will be useful, but
# WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE. See the GNU
# Affero General Public License for more details.
#
# You should have received a copy of the GNU Affero General Public
# License along with this program. If not, see
# <http://www.gnu.org/licenses/>.
#
# You can be released from the requirements of the license by purchasing
# a commercial license. Buying such a license is mandatory as soon as
# you develop activities involving the FlOpEDT/FlOpScheduler software
# without disclosing the source code of your own applications.

from base.models import Day, Course, Group, CourseType, RoomType, Module,\
    CourseStartTimeConstraint, TimeGeneralSettings, Department, UserPreference, ScheduledCourse, ModulePossibleTutors
from people.models import Tutor
from misc.assign_colors import assign_tutor_color
from MyFlOp.cosmo_specific_constraints import possible_tutors
from misc.deploy_database.deploy_database import extract_database_file
from openpyxl import load_workbook
from copy import copy
from TTapp.models import days_index



def affect_tutors():
    for m in Module.objects.all():
        mpt = ModulePossibleTutors(module=m)
        if m.abbrev == "Ct_Men":
            for t in possible_tutors['Autre']:
                mpt.possible_tutors.add(Tutor.objects.get(username=t))
        else:
            for t in possible_tutors[m.abbrev]:
                mpt.possible_tutors.add(Tutor.objects.get(username=t))
        mpt.save()
    print("Possible tutors affected")


def TGS_deploiement(department):
    T_C = TimeGeneralSettings(department=department,
                              day_start_time=60*7+30,
                              day_finish_time=25*60,
                              lunch_break_start_time=14*60,
                              lunch_break_finish_time=14*60,
                              days=[c[0] for c in Day.CHOICES])
    T_C.save()
    print("Time general settings defined")


def global_deploiement():
    extract_database_file("misc/deploy_database/Files/Cosmo/database_file_Cosmo.xlsx", 'American Cosmograph', "Cosmo")
    affect_tutors()
    TGS_deploiement(Department.objects.all()[0])
    assign_tutor_color()
    print("Tutor color assigned")
    A = Tutor(username='CHEF')
    A.save()
    A.set_password('passe')
    A.is_tutor = True
    A.is_staff = True
    A.is_superuser = True
    A.rights = 6
    A.save()
    print("Admin Tutor defined")
    deploy_preferences()


def deploy_preferences():
    C = CourseStartTimeConstraint.objects.all()[0]
    for t in Tutor.objects.all():
        ST = C.allowed_start_times
        for d in [day[0] for day in Day.CHOICES[:7]]:
            for st in ST[::2]:
                up = UserPreference(user=t, day=d, start_time=st, duration=120, valeur=8)
                if st == 24*60:
                    up.duration=60
                elif st == 8*60:
                    up.start_time=7*60+30
                    up.duration=150
                up.save()
    print("Tutor preferences defined")


def deploy_scheduled_courses(semaine_depart, bookname=None):
    if bookname is None:
        bookname = "misc/deploy_database/Files/Cosmo/COSMO-PLANNINGS-sem%ga%g-X.xlsx" % (semaine_depart, semaine_depart+4)
    salarie={
        '1': Tutor.objects.get(username='Annie'),
        '2': Tutor.objects.get(username='Jeremy'),
        '3': Tutor.objects.get(username='Frederic'),
        '4': Tutor.objects.get(username='Adelaide'),
        '5': Tutor.objects.get(username='Elsa'),
        '6': Tutor.objects.get(username='Emilie'),
        '7': Tutor.objects.get(username='Pierre'),
        '8': Tutor.objects.get(username='Marine'),
        '9': Tutor.objects.get(username='Leandre'),
        '10': Tutor.objects.get(username='Florian'),
        '11': Tutor.objects.get(username='Nicolas'),
        '12': Tutor.objects.get(username='Manon'),
        'X': None,
        'x': None
    }
    liste_postes = [Module.objects.get(abbrev="Proj"), Module.objects.get(abbrev="Caisse"),
                    Module.objects.get(abbrev="Ct_Men"), Module.objects.get(abbrev="Autre")]
    liste_groupes = [Group.objects.get(nom="Proj"), Group.objects.get(nom="Caisse"),
                    Group.objects.get(nom="Ct_Men"), Group.objects.get(nom="Autre")]
    liste_types_de_salles = [RoomType.objects.get(name="Proj"), RoomType.objects.get(name="Caisse"),
                             RoomType.objects.get(name="Ct_Men"), RoomType.objects.get(name="Autre")]
    heure=CourseType.objects.get(name='OS')
    demie_heure = CourseType.objects.get(name='OS-30')
    book = load_workbook(filename=bookname, data_only=True)
    for sheet in book:
        week = int(sheet.title[3:])+semaine_depart-1
        print(week)
        for row in range(2, 36):
            num = (row-2) % 5
            if num == 4:
                continue
            jour = Day.CHOICES[(row-2)//5][0]
            poste = liste_postes[num]
            groupe = liste_groupes[num]
            type_de_salle = liste_types_de_salles[num]
            sept_trente=str(sheet.cell(row=row, column=4).value)
            if sept_trente in salarie:
                c = Course(tutor=salarie[sept_trente], semaine=week, an=2019, type=demie_heure,
                           module=poste, groupe=groupe, room_type=type_de_salle)
                c.save()
                sc = ScheduledCourse(cours=c, day=jour, start_time=7 * 60 + 30,
                                     tutor=salarie[sept_trente])
                sc.save()
            for col in range(5,38,2):
                valeur_pile = str(sheet.cell(row=row, column=col).value)
                valeur_demie = str(sheet.cell(row=row, column=col+1).value)
                if valeur_pile == valeur_demie and valeur_pile in salarie:
                    c = Course(tutor=salarie[valeur_pile], semaine=week, an=2019, type=heure,
                               module=poste, groupe=groupe, room_type=type_de_salle)
                    c.save()
                    sc = ScheduledCourse(cours=c, day=jour, start_time=7*60+(col-3)*30,
                                         tutor=salarie[valeur_pile])
                    sc.save()
                else:
                    for valeur in [valeur_pile, valeur_demie]:
                        if valeur in salarie:
                            c = Course(tutor=salarie[valeur], semaine=week, an=2019, type=demie_heure,
                                       module=poste, groupe=groupe, room_type=type_de_salle)
                            c.save()
                            sc = ScheduledCourse(cours=c, day=jour, start_time=7*60 + (col - 3) * 30,
                                                 tutor=salarie[valeur])
                            if valeur == valeur_demie:
                                sc.start_time += 30
                            sc.save()
    print("Cours déployés!")


def copy_cell(sheet, coord):
    return {'v': sheet[coord].value, 's': copy(sheet[coord]._style)}


def convert_to_coord(scheduled_course):
    module, day, start_time = scheduled_course.cours.module, scheduled_course.day, scheduled_course.start_time
    row = 2 + 5 * days_index[day]
    for i,m in enumerate(['Proj', 'Caisse', 'Ct_Men', 'Autre']):
        if module.abbrev == m:
            row += i
            continue
    column = (start_time - (7*60+30))//30 + 4
    if scheduled_course.cours.type.duration==30:
        return [(row, column)]
    else:
        return [(row, column), (row, column+1)]


def export_scheduled_courses(semaines, work_copy=0,
                             bookname="misc/deploy_database/Files/Cosmo/Cosmo-plannings-vide.xlsx"):
    try:
        semaines = list(semaines)
    except TypeError:
        semaines = [semaines]
    book = load_workbook(filename=bookname)
    sheet = book['Sem']
    salaries = {s: copy_cell(sheet, 'AM%d'%i)
                    for s,i in list(zip(
                ["Annie", "Jeremy", "Frederic", "Adelaide", "Elsa",
                "Emilie", "Pierre", "Marine", "Oceane", "Leandre",
                 "Florian", "Nicolas", "Manon", None],
                [3,5,8,10,13,
                 15,18,20,23,23,
                 25,28,30, 33]))}
    depart = 43464
    for semaine in semaines:
        new_sheet = book.copy_worksheet(sheet)
        new_sheet.title = 'Sem%d' % semaine
        new_sheet['A1'].value = 'Sem%d' % semaine
        new_sheet['B4'].value = depart + (semaine - 1) * 7
        for sc in ScheduledCourse.objects.filter(cours__semaine=semaine, copie_travail=work_copy):
            try:
                salarie = sc.tutor.username
            except AttributeError:
                salarie = None
            for row, column in convert_to_coord(sc):
                a = new_sheet.cell(row=row, column=column)
                a.value = salaries[salarie]['v']
                a._style = salaries[salarie]['s']
    book.remove(sheet)
    bookname = "misc/deploy_database/Files/Cosmo/Cosmo-plannings-generes-semaines-%g-a-%g.xlsx" % (semaines[0], semaines[-1])
    book.save(bookname)
    print("Fichier excel sauvé à l'adresse %s" % bookname)
    return book



