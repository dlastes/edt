# coding: utf-8
# !/usr/bin/python

# This file is part of the FlOpEDT/FlOpScheduler project.
# Copyright (c) 2017
# Authors: Iulian Ober, Paul Renaud-Goud, Pablo Seban, et al.
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU Affero General Public License as
# published by the Free Software Foundation, either version 3 of the
# License, or (at your option) any later version.
#
# This program is distributed in the hope that it will be useful, but
# WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE. See the GNU
# Affero General Public License for more details.
#
# You should have received a copy of the GNU Affero General Public
# License along with this program. If not, see
# <http://www.gnu.org/licenses/>.
#
# You can be released from the requirements of the license by purchasing
# a commercial license. Buying such a license is mandatory as soon as
# you develop activities involving the FlOpEDT/FlOpScheduler software
# without disclosing the source code of your own applications.

from openpyxl import load_workbook

from base.weeks import annee_courante

from base.models import Course, Group, Module, GroupType, Period, CourseType, RoomType

from people.models import FullStaff, SupplyStaff, Tutor

from misc.assign_module_color import assign_color

bookname='misc/deploy_database/planif_file.xlsx'

def deploy_planif_file(bookname=bookname):

    book = load_workbook(filename=bookname)

    for sheet in book:

        ################ Writing weeks in sheet 2 ################

        period_name = sheet.title

        period = Period.objects.get(name=period_name)
        module_col = 1
        course_type_col = 2
        group_type_col = 3
        week_row = 3
        cell_row = 4

        course_type_name = sheet.cell(row=cell_row, column=course_type_col).value

        while course_type_name is not None:
            if course_type_name == 'Not None':
                cell_row += 1
                course_type_name = sheet.cell(row=cell_row, column=course_type_col).value
                continue

            course_type = CourseType.objects.get(name=course_type_name)

            module_name = sheet.cell(row=cell_row, column=module_col).value
            if module_name is not None:
                module = Module.objects.get(abbrev=module_name, period=period)
                train_prog = module.train_prog

            group_type_name = sheet.cell(row=cell_row, column=group_type_col).value

            cell_col = 4
            week = sheet.cell(row=week_row, column=cell_col).value

            while week is not None:
                week = int(week)
                courses_nb = sheet.cell(row=cell_row, column=cell_col).value
                if courses_nb is not None:
                    if week > 30:
                        year = annee_courante
                    else:
                        year = annee_courante + 1
                    for group in Group.objects.filter(type__name=group_type_name, train_prog=train_prog):
############################ We must define a room_type for each course!
                        for _ in range(courses_nb):
                            R = RoomType.objects.all()
                            room_type = R[0]
############################ We must define a tutor for each course!
                            c = Course(semaine=week, an=year, groupe=group, type=course_type,
                                       module=module, tutor=module.head, room_type=room_type)
                            c.save()
                cell_col += 1
                week = sheet.cell(row=week_row, column=cell_col).value
            cell_row += 1
            course_type_name = sheet.cell(row=cell_row, column=course_type_col).value

    assign_color()
