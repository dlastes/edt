from csv import DictWriter, writer
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

from django.conf import settings

from base.weeks import year_by_week

DAYS = [('monday', 'lundi'),
                 ('tuesday', 'mardi'),
                 ('wednesday', 'mercredi'),
                 ('thursday', 'jeudi'),
                 ('friday', 'vendredi')]

def calc_start_time(time):
    t = time.split(':')
    return int(t[0])*60 + int(t[1])


def translate_day(day):
    x = filter(lambda x, y: y.uppercase() == day.uppercase(), DAYS)
    return [x for x, y in DAYS if y.upper() == day.upper()][0]


target_path = f"{settings.MEDIA_ROOT}/importation/dispo_file_.csv"


def convert_xlsx2csv(source_path, target_path=target_path):
    NAME_COL = 1
    USERNAME_COL = 2
    NB_COURSES = 3
    START_WEEK_COL = 4
    END_WEEK_COL = 23
    SUM_COL = 24
    SLOT_ROW = 4
    DAY_ROW = 2
    DURATION = 120
    with open(target_path, 'w') as csvfile:
        fieldnames = ['prof', 'year', 'week', 'day', 'start_time', 'duration', 'value']
        csv_writer = writer(csvfile, delimiter=',')
        csv_writer.writerow([g for g in fieldnames])
        wb = load_workbook(filename=source_path)
        for sheet in wb.worksheets:
            rank = 5
            week = sheet.title.split(' ')[1]
            while sheet.cell(row=rank, column=NAME_COL).value is not None:
                for col in range(START_WEEK_COL, END_WEEK_COL+1):
                    prof = sheet.cell(row=rank, column=USERNAME_COL).value
                    year = year_by_week(int(week))
                    if not isinstance(sheet.cell(row=DAY_ROW, column=col), MergedCell):
                        day = translate_day(sheet.cell(row=DAY_ROW, column=col).value)
                    time = calc_start_time(sheet.cell(row=SLOT_ROW, column=col).value)
                    value = sheet.cell(row=rank, column=col).value
                    print(prof, year, week, day, time, DURATION, value)
                    csv_writer.writerow([prof, year, week, day, time, DURATION, value])
                rank += 1
    return target_path
