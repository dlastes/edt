from openpyxl import load_workbook
from copy import copy

from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.worksheet.datavalidation import DataValidation

from configuration.make_planif_file import append_row
from django.db.models import Count

from django.conf import settings

from base.models import Course
from people.models import User

from exportation.layout_calendar import get_date_start
from base.weeks import year_by_week


# def copy_cell(source_cell, coord, tgt):
#     tgt[coord].value = source_cell.value
#     if source_cell.has_style:
#         tgt[coord]._style = copy(source_cell._style)
#     return tgt[coord]


empty_bookname=f"{settings.MEDIA_ROOT}/importation/base/empty_dispo_file.xlsx"
target_bookname=f"{settings.MEDIA_ROOT}/importation/base/dispo_file_s.xlsx"


def make_dispo_file(periode, empty_bookname=empty_bookname, target_bookname=target_bookname):
    new_book = load_workbook(filename=empty_bookname)
    empty_rows = list(new_book['rows'].rows)
    print(len(empty_rows))
    WEEK_COL = 3
    NAME_COL = 1
    USERNAME_COL = 2
    NB_COURSES = 3
    SUM_COL = 24
    START_SLOT = 4
    END_SLOT = 23

    for week in range(periode.starting_week, periode.ending_week+1):
        # sheet = new_book.create_sheet(f"Semaine {week}")
        print(new_book.active)
        # new_book.get_sheet_by_name
        sheet = new_book.copy_worksheet(new_book['empty'])
        sheet.title = f"Semaine {week}"

        sheet.cell(row=1, column=WEEK_COL).value = f"Semaine {week}"
        for day in range(1, 6):
            date = get_date_start(year_by_week(week), week, day)
            print(date.strftime("%d/%m"))
            sheet.cell(row=3, column=(START_SLOT * day)).value = date.strftime("%d/%m")

        # DataValidation
        dv = DataValidation(type="decimal",
                            operator="between",
                            formula1=0,
                            formula2=4)
        dv.error = "Le poids de la disponibilité doit être compris entre 0 et 4."
        dv.errorTitle = "Entrée invalide"

        sheet.add_data_validation(dv)

        # Nb de cours par tutor
        tutors_courses = Course.objects.filter(week=week).values('tutor_id').annotate(Count('id'))
        print(tutors_courses)
        color = 0
        rank = 5
        for tutor_course in tutors_courses:
            color = color % 5
            append_row(sheet, empty_rows, color + 1, rank, 24)
            tutor = User.objects.get(id=tutor_course['tutor_id'])
            sheet.cell(row=rank, column=NAME_COL).value = f"{tutor.last_name} {tutor.first_name}"
            sheet.cell(row=rank, column=USERNAME_COL).value = tutor.username

            sheet.cell(row=rank, column=NB_COURSES).value = tutor_course['id__count']
            sheet.cell(row=rank, column=SUM_COL).value = f"=COUNTIF(D{rank}:W{rank},\">0\")"
            for col in range(START_SLOT, END_SLOT+1):
                dv.add(sheet.cell(row=rank, column=col))
            color += 1
            rank += 1

        sheet.conditional_formatting.add(f"D5:W{rank-1}",
                                         ColorScaleRule(start_type='min', start_value=0, start_color='FF0000',
                                                        mid_type='num', mid_value=2, mid_color='FFD700',
                                                        end_type='max', end_value=4, end_color='00AA00'))
        rank += 1
        append_row(sheet, empty_rows, 7, rank, 1)
        sheet.merge_cells(start_row=rank, end_row=rank, start_column=1, end_column=2)

        for row in range(0, 5):
            rank += 1
            append_row(sheet, empty_rows, 8 + row, rank, 3)

    new_book.remove(new_book['rows'])
    new_book.remove(new_book['empty'])
    new_book.save(filename=target_bookname)
    return target_bookname

