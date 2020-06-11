from openpyxl import load_workbook

import logging
logger = logging.getLogger(__name__)

people_sheet = 'Intervenants'
rooms_sheet = 'Salles'
groups_sheet = 'Groupes'
modules_sheet = 'Modules'
cours_sheet = 'Cours'
settings_sheet = 'Paramètres'

REASONABLE = 3141 # large enough?

#################################################
#                                               #
#   Helper functions to parse various types     #
#   of cells, either individually or by range   #
#                                               #
#################################################

def parse_time(sheet, row, column):
    "Helper function to get a time out of a cell"
    "as a number of minutes since midnight"
    "(will return None if anything goes wrong)"
    try:
        val = sheet.cell(row=row, column=column).value
        return 60 * val.hour + val.minute
    except:
        return None

def parse_time_list_in_line(sheet, row, col_start):
    "Parse a line representing a list of times"
    "(stop at the first empty cell)"

    result = []
    col = col_start
    while col < REASONABLE:
        val = parse_time(sheet, row, col)
        if val == None:
            break
        result.append(val)
        col = col + 1

    return result

def parse_string(sheet, row, column):
    "Helper function to get a clean string out of a cell"
    "(will return '' if there's nothing to see)"

    val = sheet.cell(row=row, column=column).value
    if val == None:
        val=''
    val = str(val).strip()

    return val

def parse_string_list_in_line(sheet, row, col_start):
    "Parse a line representing a list of strings"
    "(stop at the first empty cell)"

    result = []
    col = col_start
    while col < REASONABLE:
        val = parse_string(sheet, row, col)
        if val == '':
            break
        result.append(val)
        col = col + 1

    return result

def parse_string_list_dictionary(sheet, row_start, col_start, row_end = REASONABLE):
    "Parse a block, turning it into a dictionary of string lists"
    "The first column gives the keys, the line at the right of the key is"
    "the associated value"
    result = dict()
    row = row_start
    while row < row_end:
        name = parse_string(sheet, row, col_start)
        if name == '':
            row = row + 1
            continue
        if name in result:
            logger.warning(f"At row {row:d}, key {name} is a duplicate: ignoring the line")
            row = row + 1
            continue
        result[name] = parse_string_list_in_line(sheet, row, col_start + 1)
        row = row + 1

    return result

################################
#                              #
#   Various helper functions   #
#                              #
################################


def find_cell(sheet, marker, row = 1, col = 1):
    "Helper function to find the marker of a data block"
    "(Will return either row, col or None, None)"
    "(With optional parameters to start the search at some position)"

    while row < REASONABLE:
        while col < REASONABLE:
            if parse_string(sheet, row, col) == marker:
                return row, col
            col = col + 1
        row = row + 1
        col = 1
    logger.warning(f"The marker cell {marker} wasn't found")
    return None, None

#################################################
#                                               #
#   Parser functions for the different pages    #
#                                               #
#################################################

def parse_rooms(sheet):
    row_groups, col_groups = find_cell(sheet, 'Groupes')
    row_cats, col_cats = find_cell(sheet, 'Catégories')
    if col_groups != col_cats or row_cats < row_groups:
        logger.warning(f"The marker cells in sheet {rooms_sheet} are misplaced")
        return set(), dict(), dict()

    #
    # parse the groups
    #
    groups = dict()
    pre_groups = parse_string_list_dictionary(sheet, row_groups + 1, col_groups, row_cats)

    # drop empty groups
    empty = set() # empty groups
    for name, lst in pre_groups.items():
        if len(lst) == 0:
            empty.add(name)
    if len(empty) > 0:
        logger.warning("Some groups of rooms are empty, hence ignored: {0:s}".format(', '.join(empty)))
        for solo in empty:
            del pre_groups[solo]

    # don't accept group names in groups
    group_names = pre_groups.keys()
    for name, lst in pre_groups.items():
        bad = set.intersection(set(lst), group_names)
        if len(bad) > 0:
            logger.warning("Group '{0:s}' contains group names, ignoring them: {1:s}".format(name,', '.join(bad)))
            groups[name] = set.difference(set(lst), group_names)
        else:
            groups[name] = set(lst)

    #
    # parse the categories
    #
    categories = dict()
    pre_cats = parse_string_list_dictionary(sheet, row_cats + 1, col_cats)

    # drop empty categories
    empty = set() # empty groups
    for name, lst in pre_cats.items():
        if len(lst) == 0:
            empty.add(name)
    if len(empty) > 0:
        logger.warning("Some categories of rooms are empty, hence ignored: {0:s}".format(', '.join(empty)))
        for solo in empty:
            del pre_cats[solo]

    # don't accept category names in categories
    cat_names = pre_cats.keys()
    for name, lst in pre_cats.items():
        bad = set.intersection(set(lst), cat_names)
        if len(bad) > 0:
            logger.warning("Category '{0:s}' contains category names, ignoring them: {1:s}".format(name,', '.join(bad)))
            categories[name] = set.difference(set(lst), cat_names)
        else:
            categories[name] = set(lst)

    #
    # Build the set of rooms
    #
    rooms = set()
    for lst in groups.values():
        rooms.update(lst)
    for lst in categories.values():
        rooms.update(lst)
    rooms.difference_update(group_names)
    rooms.difference_update(cat_names)

    return rooms, groups, categories

def parse_people(sheet):
    row, col = find_cell(sheet, "Identifiant")
    if row == None:
        logger.warning(f"The marker cell in sheet {people_sheet} is missing")
        return dict()

    row = row + 1
    result = dict()
    while row < REASONABLE:
        id_ = parse_string(sheet, row, col)
        if id_ == '':
            row = row + 1
            continue
        if id_ in result:
            logger.warning(f"Duplicate identifier '{id_}' in row {row}: ignoring the line")
            row = row + 1
            continue
        result[id_] = {'first_name': parse_string(sheet, row, col + 1),
                       'last_name': parse_string(sheet, row, col + 2),
                       'email': parse_string(sheet, row, col + 3),
                       'status': parse_string(sheet, row, col + 4),
                       'employer': parse_string(sheet, row, col + 5)}
        row = row + 1
    return result

def parse_modules(sheet):
    row, col = find_cell(sheet, "Abréviation")
    if row == None:
        logger.warning(f"The marker cell in sheet {modules_sheet} is missing")
        return dict()

    row = row + 1
    result = dict()
    while row < REASONABLE:
        id_ = parse_string(sheet, row, col)
        if id_ == '':
            row = row + 1
            continue
        if id_ in result:
            logger.warning(f"Duplicate module identifier '{id_}' in row {row}: ignoring the line")
            row = row + 1
            continue
        result[id_] = {'PPN': parse_string(sheet, row, col + 1),
                       'name': parse_string(sheet, row, col + 2),
                       'promotion': parse_string(sheet, row, col + 3),
                       'period': parse_string(sheet, row, col + 4),
                       'responsable': parse_string(sheet, row, col + 5)}
        row = row + 1
    return result

def parse_cours(sheet):
    row, col = find_cell(sheet, 'Type')
    if row == None:
        logger.warning(f"The marker cell in sheet {cours_sheet} is missing")
        return dict()

    row = row + 1
    result = dict()
    while row < REASONABLE:
        id_ = parse_string(sheet, row, col)
        if id_ == '':
            row = row + 1
            continue
        if id_ in result:
            logger.warning(f"Duplicate course identifier '{id_}' in row {row}: ignoring the line")
            row = row + 1
            continue
        try:
            duree = int(parse_string(sheet, row, col + 1))
        except:
            logger.warning(f"Invalid duration in row {row} and column {col + 1}: ignoring the line")
            row = row + 1
            continue
        result[id_] = {'duration': duree,
                       'group_types': set(parse_string_list_in_line(sheet, row, col + 2)),
                       'start_times': set(parse_time_list_in_line(sheet, row + 1, col + 2))}
        row = row + 2 # sic
    return result

def parse_settings(sheet):
    jalons = dict()
    row, col = find_cell(sheet, 'Jalon')
    if row == None:
        logger.warning(f"The 'Jalon' cell in sheet {settings_sheet} is missing : using defaults")
        jalons['day_start_time'] = 0 * 60
        jalons['day_finish_time'] = 23 * 60
        jalons['lunch_break_start_time'] = 12 * 60
        jalons['lunch_break_finish_time'] = 13 * 60
    else:
        val = parse_time(sheet, row + 1, col + 1)
        if val == None:
            logger.warning("Invalid time for day start time")
            val = 0 * 60
        jalons['day_start_time'] = val
        val = parse_time(sheet, row + 2, col + 1)
        if val == None:
            logger.warning("Invalid time for day finish time")
            val = 23 * 60
        jalons['day_finish_time'] = val
        val = parse_time(sheet, row + 3, col + 1)
        if val == None:
            logger.warning("Invalid time for lunch break start time")
            val = 12 * 60
        jalons['lunch_break_start_time'] = val
        val = parse_time(sheet, row + 4, col + 1)
        if val == None:
            logger.warning("Invalid time for lunch break finish time")
            val = 13 * 60
        jalons['lunch_break_finish_time'] = val

    row, col = find_cell(sheet, 'Granularité')
    duration = 60
    if row == None:
        logger.warning(f"The 'Granularité' cell in sheet {params_sheet} is missing : using default 60")
    else:
        try:
            duration = int(parse_string(sheet, row, col + 1))
        except:
            logger.warning(f"The 'Granularité' in sheet {params_sheet} has an invalid value: using default 60")

    days = set()
    row, col = find_cell(sheet, 'Jours ouvrables')
    if row == None:
        logger.warning(f"The 'Jours ouvrables' cell in sheet {params_sheet} is missing")
    else:
        # FIXME base.timing.Day has a CHOICES with this, but it's not available here
        for index, day in enumerate(['m', 'tu', 'w', 'th', 'f', 'sa', 'su']):
            if parse_string(sheet, row + 2, col + index) == 'X':
                days.add(day)

    periods = dict()
    row, col = find_cell(sheet, 'Périodes')
    if row == None:
        logger.warning(f"The 'Périodes' cell in sheet {params_sheet} is missing")
    else:
        row = row + 2
        while row < REASONABLE:
            id_ = parse_string(sheet, row, col)
            if id_ == '':
                row = row + 1
                continue
            if id_ in periods:
                logger.warning(f"Duplicate period identifier '{id_}' in row {row}: ignoring the line")
                row = row + 1
                continue
            periods[id_] = (parse_string(sheet, row, col + 1), parse_string(sheet, row, col + 2))
            row = row + 1

    return {'jalons': jalons,
            'default_preference_duration': duration,
            'days': days,
            'periods': periods }

def parse_groups(sheet):
    row_prom, col_prom = find_cell(sheet, 'Identifiant')
    if row_prom == None:
        logger.warning(f"The 'Identifiant' cell in sheet {groups_sheet} is missing")
        return dict(), set(), dict()

    row_nat, col_nat = find_cell(sheet, 'Identifiant', row_prom + 1)
    if row_nat == None:
        logger.warning(f"The second 'Identifiant' cell in sheet {groups_sheet} is missing")
        return dict(), set(), dict()

    row_grp, col_grp = find_cell(sheet, 'Identifiant', row_nat + 1)
    if row_grp == None:
        logger.warning(f"The third 'Identifiant' cell in sheet {groups_sheet} is missing")
        return dict(), set(), dict()

    promotions = dict()
    row = row_prom + 1
    while row < row_nat: # should stop before

        id_ = parse_string(sheet, row, col_prom)
        if id_ == '':
            break
        if id_ in promotions:
            logger.warning(f"Duplicate promotion identifier '{id_}' in row {row}: ignoring the line")
            row = row + 1
            continue
        promotions[id_] = parse_string(sheet, row, col_prom + 1)
        row = row + 1

    group_types = set()
    row = row_nat + 1
    while row < row_grp: # should stop before

        id_ = parse_string(sheet, row, col_nat)
        if id_ == '':
            break
        if id_ in group_types:
            logger.warning(f"Duplicate group type identifier '{id_}' in row {row}")
        group_types.add(id_)
        row = row + 1

    groups = dict()
    row = row_grp + 1
    while row < REASONABLE:

        id_ = parse_string(sheet, row, col_grp)
        if id_ == '':
            row = row + 1
            continue
        if id_ in groups:
            logger.warning(f"Duplicate group identifier '{id_}' in row {row}: ignoring the line")
            row = row + 1
            continue
        promotion = parse_string(sheet, row, col_grp + 1)
        if not promotion in promotions:
            logger.warning(f"Group in non-existing promotion '{promotion}' at row {row}: ignoring the line")
            row = row + 1
            continue
        nature = parse_string(sheet, row, col_grp + 2)
        if not nature in group_types:
            logger.warning(f"Group in non-existing nature '{nature}' at row {row}: ignoring the line")
            row = row + 1
            continue
        parent_ = parse_string(sheet, row, col_grp + 3)
        if parent_ == '':
            parent = set()
        else:
            parent = {parent_}
        groups[id_] = {'promotion': promotion,
                       'nature' : nature,
                       'parent' : parent}
        row = row + 1

    return promotions, group_types, groups


#################################################
#                                               #
#           Main parsing function               #
#                                               #
#################################################


def parse_file(filename = 'file_essai.xlsx'):
    try:
        wb = load_workbook(filename, data_only=True)

        sheet = wb[rooms_sheet]
        if not sheet:
            logger.warning(f"Sheet {rooms_sheet} doesn't exist")
            return None

        rooms, room_groups, room_categories = parse_rooms(sheet)

        sheet = wb[people_sheet]
        if not sheet:
            logger.warning(f"Sheet {people_sheet} doesn't exist")
            return None

        people = parse_people(sheet)

        sheet = wb[modules_sheet]
        if not sheet:
            logger.warning(f"Sheet {modules_sheet} doesn't exist")
            return None

        modules = parse_modules(sheet)

        sheet = wb[cours_sheet]
        if not sheet:
            logger.warning(f"Sheet {cours_sheet} doesn't exist")
            return None

        cours = parse_cours(sheet)

        sheet = wb[settings_sheet]
        if not sheet:
            logger.warning(f"Sheet {settings_sheet} doesn't exist")
            return None

        settings = parse_settings(sheet)

        sheet = wb[groups_sheet]
        if not sheet:
            logger.warning(f"Sheet {groups_sheet} doesn't exist")
            return None

        promotions, group_types, groups = parse_groups(sheet)

        return {'rooms' : rooms,
                'room_groups' : room_groups,
                'room_categories' : room_categories,
                'people' : people,
                'modules' : modules,
                'cours' : cours,
                'settings' : settings,
                'promotions': promotions,
                'group_types' : group_types,
                'groups' : groups }
    except FileNotFoundError as ex:
        logger.warning("Database file couldn't be opened: ", ex)
        return None

# dirty, but for testing purposes it's nice
if __name__ == '__main__':
    print("""===== WARNINGS (expected) ======\n"""
          """At row 15, key doublon is a duplicate: ignoring the line\n"""
          """Some groups of rooms are empty, hence ignored: rien\n"""
          """Group 'test' contains group names, ignoring them: TP_phy\n"""
          """Some categories of rooms are empty, hence ignored: vide\n"""
          """Category 'langue' contains category names, ignoring them: TP\n"""
          """Duplicate identifier 'AB' in row 7: ignoring the line\n"""
          """Duplicate module identifier 'doublon' in row 22: ignoring the line\n"""
          """Duplicate course identifier 'TP' in row 16: ignoring the line\n"""
          """Duplicate period identifier 'S3' in row 18: ignoring the line\n"""
          """Duplicate promotion identifier 'INFO1' in row 8: ignoring the line\n"""
          """Duplicate group identifier 'MP' in row 42: ignoring the line\n""")

    print("===== WARNINGS (actual) ======")
    results = parse_file()
    print("===== RESULTS ===========")
    print(results)
