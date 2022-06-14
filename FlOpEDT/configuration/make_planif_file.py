# coding: utf-8
# !/usr/bin/python

# This file is part of the FlOpEDT/FlOpScheduler project.
# Copyright (c) 2017
# Authors: Iulian Ober, Paul Renaud-Goud, Pablo Seban, et al.
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU Affero General Public License as
# published by the Free Software Foundation, either version 3 of the
# License, or (at your option) any later version.
#
# This program is distributed in the hope that it will be useful, but
# WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE. See the GNU
# Affero General Public License for more details.
#
# You should have received a copy of the GNU Affero General Public
# License along with this program. If not, see
# <http://www.gnu.org/licenses/>.
#
# You can be released from the requirements of the license by purchasing
# a commercial license. Buying such a license is mandatory as soon as
# you develop activities involving the FlOpEDT/FlOpScheduler software
# without disclosing the source code of your own applications.

import logging

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from base.models import StructuralGroup, Module, Period, CourseType, RoomType, Course, TransversalGroup
from people.models import Tutor

from copy import copy

from django.db.models import Count

from django.utils.translation import gettext_lazy as _


logger = logging.getLogger(__name__)


def as_text(value):
    if value is None:
        return ""
    return str(value)


def column_letter(col):
    if col <= 26:
        return chr(64 + col)
    else:
        return chr(64 + (col - 1) // 26) + chr(65 + (col - 1) % 26)


def append_row(work_sheet, rows_to_append, row_number, rank, until):
    """
    Add a copy of 'rows_to_append[row_number]' to work_sheet with 'until' columns
    """
    row = rows_to_append[row_number - 1]
    for cell in row[:until]:
        new_cell = work_sheet.cell(row=rank, column=cell.col_idx, value=cell.value)
        if cell.has_style:
            new_cell.font = copy(cell.font)
            new_cell.border = copy(cell.border)
            new_cell.fill = copy(cell.fill)
            new_cell.number_format = copy(cell.number_format)
            new_cell.protection = copy(cell.protection)
            new_cell.alignment = copy(cell.alignment)

def order_CT(department):
    CT = []
    CT += list(CourseType.objects.filter(department=department, name__contains='CM'))
    CT += list(CourseType.objects.filter(department=department, name__contains='A'))
    CT += list(CourseType.objects.filter(department=department, name__contains='TD'))
    CT += list(CourseType.objects.filter(department=department, name__contains='TP'))
    CT += CourseType.objects.filter(department=department).exclude(name__contains='TP')\
        .exclude(name__contains='A').exclude(name__contains='TD').exclude(name__contains='CM')
    return CT


default_empty_bookname = 'media/configuration/empty_planif_file.xlsx'

def adjust_column_length(sheet):
    for i, col in enumerate(sheet.columns):
        length = len(as_text(col[0].value))
        if length == 2:
            length = 1
        adjusted_length = (length + 2) * 1.2
        sheet.column_dimensions[get_column_letter(i + 1)].width = adjusted_length


def make_planif_file(department, empty_bookname=default_empty_bookname, target_repo="media/configuration",
                     with_courses=False):
    new_book = load_workbook(filename=empty_bookname)

    # Define the list of possible tutors and possible room_types
    rules = new_book['Rules']
    tutor_row = 7
    tutor_col = 2
    tutor_list = list(t.username for t in Tutor.objects.filter(departments=department))
    tutor_list.sort()
    for tutor in tutor_list:
        rules.cell(row=tutor_row, column=tutor_col).value = tutor
        tutor_col += 1

    # get the last tutor letter for the validator
    # (and let room for 5 more tutors....)
    last_tutor_letter = get_column_letter(tutor_col + 5)
    tutor_validator = DataValidation(type="list", formula1=f"Rules!$B$7:${last_tutor_letter}$7", allow_blank=True)
    tutor_validator.error = "Ce prof n'est pas dans la liste de l'onglet Rules"
    tutor_validator.errorTitle = 'Erreur de prof'
    tutor_validator.prompt = 'Choisir un prof dans la liste'
    tutor_validator.promptTitle = 'Prof possibles'

    room_type_row = 12
    room_type_col = 2
    room_type_list = list(rt.name for rt in RoomType.objects.filter(department=department))
    room_type_list.sort()
    for rt in room_type_list:
        rules.cell(row=room_type_row, column=room_type_col).value = rt
        room_type_col += 1

    # get the last room_type letter for the validator
    # (and let room for 5 more tutors....)
    last_room_type_letter = get_column_letter(room_type_col + 5)

    room_type_validator = DataValidation(type="list", formula1=f"Rules!$B$12:${last_room_type_letter}$12",
                                         allow_blank=True)
    room_type_validator.error = "Ce type de salle n'est pas dans la liste de l'onglet Rules"
    room_type_validator.errorTitle = 'Erreur de type de salle'
    room_type_validator.prompt = 'Choisir un type de salle dans la liste'
    room_type_validator.promptTitle = 'Type de salles possibles'


    empty_rows = list(new_book['empty'].rows)
    recap_rows = list(new_book['empty_recap'].rows)
    new_book.create_sheet('Recap')
    last_row = {}
    last_column_letter = {}
    first_column_letter = {}
    if department.abbrev in ['INFO', 'RT', 'GIM', 'CS']:
        CT = order_CT(department)
    else:
        CT = CourseType.objects.filter(department=department)
    # We go through each period and create a sheet for each period
    for p in Period.objects.filter(department=department):
        logger.info(p)
        new_book.create_sheet(p.name)
        sheet = new_book[p.name]
        sheet.add_data_validation(tutor_validator)
        sheet.add_data_validation(room_type_validator)
        ################ Writing line 1 with weeks ################
        week_col_dict = {}
        rank = 1
        FIRST_WEEK_COL = 8
        first_column_letter[p] = column_letter(FIRST_WEEK_COL)
        week_col = FIRST_WEEK_COL

        if p.starting_week < p.ending_week:
            weeks = p.ending_week - p.starting_week + 1
            cols = weeks + 8
            append_row(sheet, empty_rows, 1, rank, cols)
            for i in range(p.starting_week, p.ending_week + 1):
                week_col_dict[i] = week_col
                sheet.cell(row=rank, column=week_col).value = i
                week_col += 1
            VERIF_COL = week_col

        else:

            weeks = (53 - p.starting_week) + p.ending_week
            cols = weeks + 8
            append_row(sheet, empty_rows, 1, rank, cols)
            for i in range(p.starting_week, 53):
                week_col_dict[i] = week_col
                sheet.cell(row=rank, column=week_col).value = i
                week_col += 1

            for i in range(1, p.ending_week + 1):
                week_col_dict[i] = week_col
                sheet.cell(row=rank, column=week_col).value = i
                week_col += 1
            VERIF_COL = week_col
        sheet.cell(row=rank, column=week_col).value = "VERIF"
        rank += 1
        append_row(sheet, empty_rows, 5, rank, cols)
        sheet.cell(row=rank, column=7).value = "MAX"
        rank += 1
        append_row(sheet, empty_rows, 5, rank, cols)
        rank += 1
        c = sheet.cell(row=rank, column=8)
        sheet.freeze_panes = c
        last_column_letter[p] = column_letter(cols-1)
        append_row(sheet, empty_rows, 4, rank, cols)
        rank += 1
        first_line = rank

        ################ A line per module per CourseType ################
        for mod in Module.objects.filter(period=p):
            courses = Course.objects.filter(module=mod)
            logger.info(f"Module {mod}")
            for ct in CT:
                type_courses = courses.filter(type=ct)
                append_row(sheet, empty_rows, 2, rank, cols)
                sheet.cell(row=rank, column=1).value = mod.abbrev
                sheet.cell(row=rank, column=2).value = '=$C%d&"_"&$E%d' % (rank, rank)
                sheet.cell(row=rank, column=3).value = ct.name
                sheet.cell(row=rank, column=4).value = ct.duration
                sheet.cell(row=rank, column=5).value = 'Prof'
                sheet.cell(row=rank, column=6).value = 'Type de Salle'
                sheet.cell(row=rank, column=7).value = 'Groupes'
                sheet.cell(row=rank, column=VERIF_COL).value = '=SUM(%s%d:%s%d)' % (first_column_letter[p], rank,
                                                                                    last_column_letter[p], rank)
                rank += 1
                groups = set(StructuralGroup.objects.filter(train_prog=mod.train_prog,
                                                            type__in=ct.group_types.all())) \
                         | set(TransversalGroup.objects.filter(train_prog=mod.train_prog,
                                                               type__in=ct.group_types.all()))
                if with_courses:
                    for c in type_courses.distinct('groups'):
                        groups |= set(c.groups.all())

                nb_groups = len(groups)
                if nb_groups:
                    if with_courses:
                        relevant_groups_dict = {}
                        for c in type_courses:
                            relevant_groups = c.groups.all()
                            group_to_be_written = ';'.join(g.name for g in relevant_groups)
                            relevant_groups_dict[group_to_be_written] = c.groups.all()
                        for groups_name, groups in relevant_groups_dict.items():
                            # This 3 lines code allow to limit the courses to those which have
                            # exactly groups as groups...
                            group_courses = type_courses.annotate(count=Count('groups')).filter(count=groups.count())
                            for gp in groups:
                                group_courses = group_courses.filter(groups=gp)
                            if not group_courses.exists():
                                continue
                            courses_tutors = type_courses.distinct('tutor')
                            for course_tutor in courses_tutors:
                                local_tutor = course_tutor.tutor
                                tutor_group_courses = group_courses.filter(tutor=local_tutor)
                                if not tutor_group_courses.exists():
                                    continue
                                if local_tutor is None:
                                    username = ""
                                else:
                                    username = local_tutor.username
                                append_row(sheet, empty_rows, 3, rank, cols)
                                sheet.cell(row=rank, column=1).value = mod.abbrev
                                sheet.cell(row=rank, column=2).value = '=$C%d&"_"&$E%d' % (rank, rank)
                                sheet.cell(row=rank, column=3).value = ct.name
                                sheet.cell(row=rank, column=4).value = ct.duration
                                sheet.cell(row=rank, column=5).value = username
                                room_type_validator.add(sheet.cell(row=rank, column=6))
                                sheet.cell(row=rank, column=7).value = groups_name

                                courses_weeks = type_courses.distinct('week').exclude(week__nb=0)
                                for course_week in courses_weeks:
                                    local_week = course_week.week
                                    try:
                                        week_col = week_col_dict[local_week.nb]
                                    except KeyError:
                                        continue
                                    week_tutor_group_courses_nb = tutor_group_courses.filter(week=local_week).count()
                                    sheet.cell(row=rank, column=week_col).value = week_tutor_group_courses_nb
                                rank += 1

                    else:
                        for g in groups:
                            append_row(sheet, empty_rows, 3, rank, cols)
                            sheet.cell(row=rank, column=1).value = mod.abbrev
                            sheet.cell(row=rank, column=2).value = '=$C%d&"_"&$E%d' % (rank, rank)
                            sheet.cell(row=rank, column=3).value = ct.name
                            sheet.cell(row=rank, column=4).value = ct.duration
                            tutor_validator.add(sheet.cell(row=rank, column=5))
                            room_type_validator.add(sheet.cell(row=rank, column=6))
                            sheet.cell(row=rank, column=7).value = g.name
                            rank += 1
                        sheet.cell(row=rank - nb_groups, column=VERIF_COL).value = '' \
                           '=IF(SUM(%s%d:INDIRECT(ADDRESS(MATCH(G$5,G%d:G%d,0)+ROW()-2,%d)))-$%s%d*%d=0,"OK","/!\\ -> ' \
                           '"&SUM(%s%d:INDIRECT(ADDRESS(MATCH(G$5,G%d:G%d,0)+ROW()-2,%d)))-$%s%d*%d)' % \
                               (
                                   first_column_letter[p], rank - nb_groups,
                                   rank - nb_groups, rank - nb_groups + 10,
                                   VERIF_COL - 1,
                                   column_letter(VERIF_COL),
                                   rank - nb_groups - 1, nb_groups,
                                   first_column_letter[p], rank - nb_groups,
                                   rank - nb_groups, rank - nb_groups + 10,
                                   VERIF_COL - 1,
                                   column_letter(VERIF_COL),
                                   rank - nb_groups - 1, nb_groups,
                               )
                else:
                    append_row(sheet, empty_rows, 3, rank, cols)
                    sheet.cell(row=rank, column=1).value = mod.abbrev
                    sheet.cell(row=rank, column=2).value = '=$C%d&"_"&$E%d' % (rank, rank)
                    sheet.cell(row=rank, column=3).value = ct.name
                    sheet.cell(row=rank, column=4).value = ct.duration
                    tutor_validator.add(sheet.cell(row=rank, column=5))
                    room_type_validator.add(sheet.cell(row=rank, column=6))
                    rank += 1

            ################ Separating each course with a black line ################
            append_row(sheet, empty_rows, 4, rank, cols)
            rank += 1

        ############ TOTAL line ############
        ligne_finale = rank - 2
        sheet.cell(row=rank-1, column=VERIF_COL).value = 'TOTAL'
        append_row(sheet, empty_rows, 5, rank, cols)
        for week_col in range(FIRST_WEEK_COL, cols):
            cl = column_letter(week_col)
            sheet.cell(row=rank, column=week_col).value = \
                '=SUMPRODUCT((D$%d:D$%d)*(%s$%d:%s$%d)*(G$%d:G$%d="Groupes"))/60' \
                % (first_line, ligne_finale, cl, first_line, cl, ligne_finale, first_line, ligne_finale)
            sheet.cell(row=first_line-2, column=week_col).value = '=%s%d' % (cl, rank)
        sheet.cell(row=rank, column=VERIF_COL).value = '=SUM(%s%d:%s%d)' % (first_column_letter[p], rank,
                                                                            last_column_letter[p], rank)
        sheet.cell(row=first_line-2, column=VERIF_COL).value = '=%s%d' % (column_letter(VERIF_COL), rank)
        rank += 1

        ############ Other TOTAL lines ############
        rank += 1
        append_row(sheet, empty_rows, 6, rank, cols)
        sheet.cell(row=rank, column=2).value = "='Recap'!$B$1"
        sheet.cell(row=rank, column=6).value = '="TOTAL_"&$B$%d' % rank
        prof_row = rank
        for week_col in range(FIRST_WEEK_COL, cols):
            cl = column_letter(week_col)
            sheet.cell(row=rank, column=week_col).value = '=%s1' % cl
        sheet.cell(row=rank, column=VERIF_COL).value = 'TOTAL'
        #sheet.row_dimensions[rank].hidden = True
        rank += 1
        for ct in CT:
            append_row(sheet, empty_rows, 7, rank, cols)
            sheet.cell(row=rank, column=2).value = '=$F%d&"_"&$B$%d' % (rank, prof_row)
            sheet.cell(row=rank, column=6).value = ct.name
            sheet.cell(row=rank, column=7).value = '=SUM(H%d:%s%d)' % (rank, last_column_letter[p], rank)
            for week_col in range(FIRST_WEEK_COL, cols):
                cl = column_letter(week_col)
                sheet.cell(row=rank, column=week_col).value =\
                    '=SUMPRODUCT((D$%d:D$%d)*(%s$%d:%s$%d)*($B$%d:$B$%d=$B%d))/60' \
                    % (first_line, ligne_finale, cl, first_line, cl, ligne_finale, first_line, ligne_finale, rank)
            sheet.cell(row=rank, column=VERIF_COL).value = '=SUM(%s%d:%s%d)' % (first_column_letter[p], rank,
                                                                                last_column_letter[p], rank)
            #sheet.row_dimensions[rank].hidden = True
            rank += 1
        append_row(sheet, empty_rows, 8, rank, cols)
        nb_ct = len(CT) #CourseType.objects.filter(department=department).count()
        for week_col in range(FIRST_WEEK_COL-1, cols):
            cl = column_letter(week_col)
            sheet.cell(row=rank, column=week_col).value = \
                '=SUM(%s%d:%s%d)' % (cl, rank - nb_ct, cl, rank - 1)
        sheet.cell(row=rank, column=VERIF_COL).value = '=SUM(%s%d:%s%d)' % (first_column_letter[p], rank,
                                                                            last_column_letter[p], rank)
        last_row[p.name] = rank
        rank += 1

        ############ Adapting column widths ############
        adjust_column_length(sheet)
        sheet.column_dimensions['B'].hidden = True


    ############ Make recap sheet ############
    sheet = new_book['Recap']
    rank = 1
    recap_col_nb = 48
    nb_per = Period.objects.filter(department=department).count()
    append_row(sheet, recap_rows, 1, rank, recap_col_nb)
    rank += 1
    for p in Period.objects.filter(department=department):
        append_row(sheet, recap_rows, 2, rank, recap_col_nb)
        sheet.cell(row=rank, column=1).value = p.name
        sheet.cell(row=rank, column=2).value = '=SUM($C%d:AV%d)' % (rank, rank)
        for week_col in range(3, recap_col_nb + 1):
            cl = column_letter(week_col)
            sheet.cell(row=rank, column=week_col).value = \
                '=SUMPRODUCT((%s!$H$%d:$%s$%d)*(%s!$H$1:$%s$1=%s$1))' % \
                (p.name, last_row[p.name], last_column_letter[p], last_row[p.name], p.name,  last_column_letter[p], cl)
            # '=SUMIF(%s!$H$1:$%s$1;%s$1;%s!$H$%d:$%s$%d)' (p.name, last_column_letter[p], cl, p.name, last_row[p.name], last_column_letter[p], last_row[p.name])

        rank += 1
    append_row(sheet, recap_rows, 3, rank, recap_col_nb)
    for week_col in range(2, recap_col_nb+1):
        cl = column_letter(week_col)
        sheet.cell(row=rank, column=week_col).value = \
            '=SUM(%s%d:%s%d)' % (cl, rank - nb_per, cl, rank - 1)
    rank += 1


    ############ Adapting column widths ############
    adjust_column_length(sheet)


    ############ Make Assignation sheet ############
    sheet = new_book['ModuleTutorsAssignation']
    tutor_assignation_validator = DataValidation(type="list", formula1="Rules!$B$7:$EE$7", allow_blank=True)
    tutor_assignation_validator.error = "Ce prof n'est pas dans la liste de l'onglet Rules"
    tutor_assignation_validator.errorTitle = 'Erreur de prof'
    tutor_assignation_validator.prompt = 'Choisir un prof dans la liste'
    tutor_assignation_validator.promptTitle = 'Prof possibles'

    sheet.add_data_validation(tutor_assignation_validator)

    module_validator = DataValidation(type="custom", allow_blank=True)
    module_validator.prompt = "Choisir un module existant"
    sheet.add_data_validation(module_validator)


    course_type_validator = DataValidation(type="custom", allow_blank=True)
    course_type_validator.prompt = "Choisir un type de cours existant"
    sheet.add_data_validation(course_type_validator)

    for row in range(2, 100, 2):
        module_validator.add(sheet.cell(row=row, column=1))
        course_type_validator.add(sheet.cell(row=row, column=2))
        for col in range(3, 16):
            tutor_assignation_validator.add(sheet.cell(row=row, column=col))

    new_book.remove(new_book['empty_recap'])
    new_book.remove(new_book['empty'])

    filename = f'{target_repo}/planif_file_' + department.abbrev
    if with_courses:
        filename += '_with_courses'
    filename += '.xlsx'

    new_book.save(filename=filename)
