# coding: utf-8
# !/usr/bin/python

# This file is part of the FlOpEDT/FlOpScheduler project.
# Copyright (c) 2017
# Authors: Iulian Ober, Paul Renaud-Goud, Pablo Seban, et al.
# 
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU Affero General Public License as
# published by the Free Software Foundation, either version 3 of the
# License, or (at your option) any later version.
# 
# This program is distributed in the hope that it will be useful, but
# WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE. See the GNU
# Affero General Public License for more details.
# 
# You should have received a copy of the GNU Affero General Public
# License along with this program. If not, see
# <http://www.gnu.org/licenses/>.
# 
# You can be released from the requirements of the license by purchasing
# a commercial license. Buying such a license is mandatory as soon as
# you develop activities involving the FlOpEDT/FlOpScheduler software
# without disclosing the source code of your own applications.

import os
import sys
from openpyxl import *

from base.weeks import actual_year
from base.models import GenericGroup, Module, Course, CourseType, RoomType,\
    TrainingProgramme, Dependency, Period, Department, CoursePossibleTutors, ModuleTutorRepartition, CourseAdditional, \
    Week
from people.models import Tutor, UserDepartmentSettings
from people.tutor import fill_default_user_preferences
from misc.assign_colors import assign_module_color
from TTapp.models import StabilizationThroughWeeks

from django.db import transaction
from django.db.models import Q

def do_assign(module, course_type, week, book):
    already_done = ModuleTutorRepartition.objects.filter(module=module, course_type=course_type,
                                                         week=week).exists()
    if already_done:
        return

    assignation_sheet = book['ModuleTutorsAssignation']
    assign_ok = False
    for assignation_row in range(1, 100):
        if assignation_sheet.cell(row=assignation_row, column=1).value == module.abbrev \
                and assignation_sheet.cell(row=assignation_row, column=2).value == course_type.name:
            assign_ok = True
            break
    if not assign_ok:
        raise Exception(f"Rien n'est prévu pour assigner {module.abbrev} / {course_type.name}...")
    column = 3
    tutor_username = assignation_sheet.cell(row=assignation_row, column=column).value
    while tutor_username is not None:
        tutor = Tutor.objects.get(username=tutor_username)
        mtr = ModuleTutorRepartition(module=module, course_type=course_type,
                                     week=week, tutor=tutor)
        nb = assignation_sheet.cell(row=assignation_row+1, column=column).value
        if nb is not None:
            nb = int(nb)
            mtr.courses_nb = nb
        mtr.save()
        column += 1
        tutor_username = assignation_sheet.cell(row=assignation_row, column=column).value
    print(f"Assignation done for {module.abbrev} / {course_type.name}!")


@transaction.atomic
def ReadPlanifWeek(department, book, feuille, week, courses_to_stabilize=None):
    sheet = book[feuille]
    period=Period.objects.get(name=feuille, department=department)
    Course.objects.filter(type__department=department, week=week, module__period=period).delete()
    # lookup week column
    wc = 1
    for wr in [1]:
        while wc < 50:
            wc += 1
            sem = sheet.cell(row=wr, column=wc).value
            if sem == float(week.nb):
                WEEK_COL = wc
                break
    try:
        WEEK_COL += 0
    except UnboundLocalError:
        print('Pas de semaine %s en %s' % (week.nb, feuille))
        return
    print("Semaine %s de %s : colonne %g" % (week.nb, feuille, WEEK_COL))

    row = 4
    module_COL = 1
    nature_COL = 3
    duree_COL = 4
    prof_COL = 5
    salle_COL = 6
    group_COL = 7
    sumtotal = 0
    while 1:
        row += 1
        if courses_to_stabilize is not None:
            if row not in courses_to_stabilize:
                courses_to_stabilize[row] = []
        is_total = sheet.cell(row=row, column=group_COL).value
        if is_total == "TOTAL":
            # print "Sem %g de %s - TOTAL: %g"%(week.nb, feuille,sumtotal)
            break

        Cell = sheet.cell(row=row, column=WEEK_COL)
        N = Cell.value
        if N is None:
            continue

        try:
            salle = sheet.cell(row=row, column=salle_COL).value
            module = sheet.cell(row=row, column=module_COL).value
            N = float(N)
            # handle dark green lines - Vert fonce
            assert isinstance(salle, str) and salle is not None
            if salle == "Type de Salle":
                nominal = int(N)
                if N != nominal:
                    print('Valeur decimale ligne %g de %s, semaine %g : on la met a 1 !' % (row, feuille, week.nb))
                    nominal = 1
                    # le nominal est le nombre de cours par groupe (de TP ou TD)
                if Cell.comment:
                    comments = Cell.comment.text.replace(' ', '').replace('\n', '').replace(',', ';').split(';')
                else:
                    comments = []

                sumtotal += nominal

                continue
            try:
                comments = comments
            except:
                comments = []

            # handle light green lines - Vert clair
            MODULE = Module.objects.get(abbrev=module, period=period)
            PROMO = MODULE.train_prog
            nature = sheet.cell(row=row, column=nature_COL).value
            prof = sheet.cell(row=row, column=prof_COL).value
            grps = sheet.cell(row=row, column=group_COL).value
            COURSE_TYPE = CourseType.objects.get(name=nature, department=department)
            ROOMTYPE = RoomType.objects.get(name=salle, department=department)
            supp_profs = []
            possible_profs = []
            if prof is None:
                TUTOR = None
            elif prof == '*':
                TUTOR = None
                do_assign(MODULE, COURSE_TYPE, week, book)
            else:
                assert isinstance(prof, str)
                prof = prof.replace('\xa0', '').replace(' ', '')
                if '|' in prof:
                    possible_profs = prof.split("|")
                    TUTOR = None
                else:
                    profs = prof.split(";")
                    prof = profs[0]
                    TUTOR = Tutor.objects.get(username=prof)
                    supp_profs = profs[1:]

            if Cell.comment:
                local_comments = Cell.comment.text.replace('\xa0', '').replace(' ', '').replace('\n', '').replace(',', ';').split(';')
            else:
                local_comments = []

            all_comments = comments + local_comments

            if isinstance(grps, int) or isinstance(grps, float):
                grps = str(int(grps))
            if not grps:
                grps = []
            else:
                grps = grps.replace('\xa0', '').replace(' ', '').replace(',', ';').split(';')
            groups = [str(g) for g in grps]

            GROUPS = list(GenericGroup.objects.filter(name__in=groups, train_prog=PROMO))
            if not GROUPS:
                raise Exception(f"Group(s) do(es) not exist {row}, week {week.nb} of {feuille}\n")

            N=int(N)


            for i in range(N):
                C = Course(tutor=TUTOR, type=COURSE_TYPE, module=MODULE, week=week,
                           room_type=ROOMTYPE)
                C.save()
                if courses_to_stabilize is not None:
                    courses_to_stabilize[row].append(C)
                for g in GROUPS:
                    C.groups.add(g)
                C.save()
                if supp_profs != []:
                    SUPP_TUTORS = Tutor.objects.filter(username__in=supp_profs)
                    for sp in SUPP_TUTORS:
                        C.supp_tutor.add(sp)
                    C.save()
                if possible_profs != []:
                    cpt = CoursePossibleTutors(course=C)
                    cpt.save()
                    for pp in possible_profs:
                        t = Tutor.objects.get(username=pp)
                        cpt.possible_tutors.add(t)
                    cpt.save()

                for after_type in [x for x in all_comments if x[0] == 'A']:
                    try:
                        n = int(after_type[1])
                        s = 2
                    except ValueError:
                        n = 1
                        s = 1
                    course_type = after_type[s:]
                    relevant_groups = set()
                    for g in GROUPS:
                        relevant_groups |= g.ancestor_groups() | {g} | g.descendants_groups()
                    courses = Course.objects.filter(type__name=course_type, module=MODULE, week=week,
                                                    groups__in=relevant_groups)
                    for course in courses[:n]:
                        P = Dependency(course1=course, course2=C)
                        P.save()

                if 'P' in all_comments:
                    course_additional, created = CourseAdditional.objects.get_or_create(course=C)
                    course_additional.visio_preference_value = 0
                    course_additional.save()
                elif 'DI' in all_comments:
                    course_additional, created = CourseAdditional.objects.get_or_create(course=C)
                    course_additional.visio_preference_value = 8
                    course_additional.save()
                if 'E' in all_comments:
                    course_additional, created = CourseAdditional.objects.get_or_create(course=C)
                    course_additional.graded = True
                    course_additional.save()
            if 'D' in comments or 'D' in local_comments and N >= 2:
                relevant_courses = Course.objects.filter(type=COURSE_TYPE, module=MODULE, groups__in=GROUPS, week=week)
                for i in range(N//2):
                    P = Dependency(course1=relevant_courses[2*i], course2=relevant_courses[2*i+1], successive=True)
                    P.save()
            if 'ND' in comments or 'ND' in local_comments and N >= 2:
                relevant_courses = Course.objects.filter(type=COURSE_TYPE, module=MODULE, groups__in=GROUPS, week=week)
                for i in range(N-1):
                    P = Dependency(course1=relevant_courses[i], course2=relevant_courses[i+1], ND=True)
                    P.save()
        except Exception as e:
            raise Exception(f"Exception ligne {row}, semaine {week.nb} de {feuille}: {e} \n")


@transaction.atomic
def extract_period(department, book, period, stabilize_courses=False, year=actual_year,
                   from_week=None, until_week=None):

    if stabilize_courses:
        courses_to_stabilize = {}
        print("Courses will be stabilized through weeks for period", period)
    else:
        courses_to_stabilize = None
    considered_weeks = set(Week.objects.all())
    if from_week is None and until_week is None:
        Course.objects.filter(module__period=period, week=None).delete()
    if until_week is not None:
        considered_weeks = set(w for w in considered_weeks if w <= until_week)
    if from_week is not None:
        considered_weeks = set(w for w in considered_weeks if w >= from_week)
    if period.starting_week < period.ending_week:
        period_weeks = Week.objects.filter(nb__gte=period.starting_week, nb__lte=period.ending_week)
        if period.ending_week > 30:
            period_weeks = period_weeks.filter(year=year)
        else:
            period_weeks = period_weeks.filter(year=year + 1)
    else:
        period_weeks = Week.objects.filter(Q(nb__gte=period.starting_week, year=year) |
                                           Q(nb__lte=period.ending_week, year=year+1))

    considered_weeks &= set(period_weeks)
    considered_weeks = list(considered_weeks)
    considered_weeks.sort()
    for week in considered_weeks:
        ReadPlanifWeek(department, book, period.name, week, courses_to_stabilize)

    if stabilize_courses:
        for courses_list in courses_to_stabilize.values():
            if len(courses_list) < 2:
                continue
            stw = StabilizationThroughWeeks.objects.create(department=department)
            for c in courses_list:
                stw.courses.add(c)


@transaction.atomic
def extract_planif(department, bookname=None, stabilize_courses=False, year=actual_year,
                   from_week=None, until_week=None, periods=None, assign_colors=True):
    '''
    Generate the courses from bookname; the school year starts in actual_year
    '''
    if bookname is None:
        bookname = 'media/configuration/planif_file_'+department.abbrev+'.xlsx'
    book = load_workbook(filename=bookname, data_only=True)
    if periods is None:
        periods = Period.objects.filter(department=department)
    for period in periods:
        extract_period(department, book, period, stabilize_courses, year=year,
                       from_week=from_week, until_week=until_week)
    if assign_colors:
        assign_module_color(department, overwrite=True)


@transaction.atomic
def extract_planif_weeks(week_year_list, department, bookname=None, periods = None):
    if bookname is None:
        bookname = 'media/configuration/planif_file_'+department.abbrev+'.xlsx'
    book = load_workbook(filename=bookname, data_only=True)
    if periods == None:
        periods = Period.objects.filter(department=department)
    for period in periods:
        for s in week_year_list:
            w = s['week']
            y = s['year']
            ReadPlanifWeek(department, book, period.name, w, y)
