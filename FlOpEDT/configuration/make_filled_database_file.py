#!/usr/bin/env python3
# -*- coding: utf-8 -*-

# This file is part of the FlOpEDT/FlOpScheduler project.
# Copyright (c) 2017
# Authors: Iulian Ober, Paul Renaud-Goud, Pablo Seban, et al.
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU Affero General Public License as
# published by the Free Software Foundation, either version 3 of the
# License, or (at your option) any later version.
#
# This program is distributed in the hope that it will be useful, but
# WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE. See the GNU
# Affero General Public License for more details.
#
# You should have received a copy of the GNU Affero General Public
# License along with this program. If not, see
# <http://www.gnu.org/licenses/>.
#
# You can be released from the requirements of the license by purchasing
# a commercial license. Buying such a license is mandatory as soon as
# you develop activities involving the FlOpEDT/FlOpScheduler software
# without disclosing the source code of your own applications.

# This code takes a XLSX database file description and turns it with
# the main function database_description_load_xlsx_file into
# structured Python data, following the data structure described in
# database_description_checker.py.
#

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from copy import copy

import logging
logger = logging.getLogger(__name__)
from configuration.database_description_xlsx import \
    people_sheet, rooms_sheet, groups_sheet, modules_sheet, courses_sheet, settings_sheet, REASONABLE, \
    find_marker_cell, time_from_integer


from base.models import Room, RoomType, Period, TransversalGroup, StructuralGroup, TrainingProgramme, GroupType, \
    Module, CourseType
from people.models import Tutor
#################################################
#                                               #
#   Filler functions for the different pages    #
#                                               #
#################################################
max_row = 200
max_col = 20
def insert_missing_rows(sheet, row_rank, data, existing_rows_number):
    merged_cells = {r.bounds[1] : [] for r in sheet.merged_cells.ranges}
    for r in sheet.merged_cells.ranges:
        merged_cells[r.bounds[1]].append((r.bounds[0], r.bounds[2]))
    # merges as dict {row : [(start_col, end_col),(st,end)], ... }
    missing_rows_number = len(data) - existing_rows_number

    if missing_rows_number > 0:

        # create copy of lines from title_row_rang +1 to max_row some row downer
        for row in range(max_row, row_rank, -1):
            new_row = row + missing_rows_number
            if new_row in merged_cells:
                for (start_col, end_col) in merged_cells[new_row]:
                    sheet.unmerge_cells(start_row=new_row, start_column=start_col,
                                        end_row=new_row, end_column=end_col)
            for col in range(1, max_col):
                cell = sheet.cell(row, col)
                new_cell = sheet.cell(new_row, col)
                new_cell.value = cell.value
                new_cell._style = copy(cell._style)
            if row in merged_cells:
                for (start_col, end_col) in merged_cells[row]:
                    sheet.merge_cells(start_row=new_row, start_column=start_col,
                                      end_row=new_row, end_column=end_col)

        for row in range(row_rank + 1, row_rank + 1 + missing_rows_number):
            if row in merged_cells:
                for (start_col, end_col) in merged_cells[row]:
                    sheet.unmerge_cells(start_row=row, start_column=start_col,
                                        end_row=row, end_column=end_col)
            for col in range(1, max_col):
                cell = sheet.cell(row_rank + 1, col)
                new_cell = sheet.cell(row, col)
                new_cell._style = copy(cell._style)
                new_cell.value = cell.value
            if row in merged_cells:
                for (start_col, end_col) in merged_cells[row]:
                    sheet.merge_cells(start_row=row, start_column=start_col,
                                      end_row=row, end_column=end_col)


def dict_from_dept_database(department):
    database_dict = {}
    database_dict['settings'] = {}
    settings = department.timegeneralsettings
    database_dict['settings']['day_start_time'] = settings.day_start_time
    database_dict['settings']['day_finish_time'] = settings.day_finish_time
    database_dict['settings']['lunch_break_start_time'] = settings.lunch_break_start_time
    database_dict['settings']['lunch_break_finish_time'] = settings.lunch_break_finish_time


def make_filled_database_file(department, filename=None):
    wb = load_workbook('media/configuration/empty_database_file.xlsx')
    if filename is None:
        filename = f'media/configuration/database_file_{department.abbrev}.xlsx'

    sheet = wb[settings_sheet]
    row, col = find_marker_cell(sheet, 'Jalon')
    sheet.cell(row=row+1, column=col+1, value=time_from_integer(department.timegeneralsettings.day_start_time))
    sheet.cell(row=row+2, column=col+1, value=time_from_integer(department.timegeneralsettings.day_finish_time))
    sheet.cell(row=row+3, column=col+1, value=time_from_integer(department.timegeneralsettings.lunch_break_start_time))
    sheet.cell(row=row+4, column=col+1, value=time_from_integer(department.timegeneralsettings.lunch_break_finish_time))

    row, col = find_marker_cell(sheet, 'Granularité')
    sheet.cell(row=row, column=col+1, value=department.timegeneralsettings.default_preference_duration)

    row, col = find_marker_cell(sheet, 'Modes')
    mode = department.mode
    if mode.visio:
        sheet.cell(row=row+1, column=col, value="Visio")
    else:
        sheet.cell(row=row + 1, column=col, value="No Visio")
    if mode.cosmo == 0:
        sheet.cell(row=row + 2, column=col, value="Educatif")
    elif mode.cosmo == 1:
        sheet.cell(row=row + 2, column=col, value="Coop.(Poste)")
    else:
        sheet.cell(row=row + 2, column=col, value="Coop. (Salarié)")


    row, col = find_marker_cell(sheet, 'Jours ouvrables')
    days = department.timegeneralsettings.days
    cols = {'m': 0,
            'tu': 1,
            'w': 2,
            'th': 3,
            'f': 4,
            'sa': 5,
            'su': 6}
    for day, delta in cols.items():
        if day in days:
            sheet.cell(row=row+2, column=col+delta, value='X')
        else:
            sheet.cell(row=row+2, column=col+delta, value=None)

    row, col = find_marker_cell(sheet, 'Périodes')
    row = row + 1
    for period in Period.objects.filter(department=department):
        row = row + 1
        sheet.cell(row=row, column=col, value=period.name)
        sheet.cell(row=row, column=col+1, value=period.starting_week)
        sheet.cell(row=row, column=col+2, value=period.ending_week)

    sheet = wb[people_sheet]
    row, col = find_marker_cell(sheet, 'Identifiant')
    for tutor in Tutor.objects.filter(departments=department):
        row = row + 1
        sheet.cell(row=row, column=col, value=tutor.username)
        sheet.cell(row=row, column=col+1, value=tutor.last_name)
        sheet.cell(row=row, column=col+2, value=tutor.first_name)
        sheet.cell(row=row, column=col+3, value=tutor.email)

        sheet.cell(row=row, column=col+4, value=tutor.status)
        try:
            sheet.cell(row=row, column=col+5, value=tutor.supplystaff.employer)
        except:
            pass

    sheet = wb[rooms_sheet]
    
    row, col_start = find_marker_cell(sheet, 'Groupes')
    room_groups = {r for r in Room.objects.filter(departments=department) if not r.is_basic}
    insert_missing_rows(sheet, row, room_groups, 9)

    for room_group in room_groups:
        row = row + 1
        col = col_start
        sheet.cell(row=row, column=col, value=room_group.name)
        for basic_room in room_group.basic_rooms():
            col = col + 1
            sheet.cell(row=row, column=col, value=basic_room.name)


    row, col_start = find_marker_cell(sheet, 'Catégories')
    room_types = RoomType.objects.filter(department=department)
    insert_missing_rows(sheet, row, room_types, 13)
    for room_type in room_types:
        row = row + 1
        col = col_start
        sheet.cell(row=row, column=col, value=room_type.name)
        for room in room_type.members.all():
            col = col + 1
            sheet.cell(row=row, column=col, value=room.name)

    sheet = wb[groups_sheet]
    train_progs = TrainingProgramme.objects.filter(department=department)
    row, col = find_marker_cell(sheet, 'Identifiant')
    # if we have too many promotions, avoid to destroy the marker below!
    insert_missing_rows(sheet, row, train_progs, 7)
    for tp in train_progs:
        row = row + 1
        sheet.cell(row=row, column=col, value=tp.abbrev)
        sheet.cell(row=row, column=col+1, value=tp.name)

    wb.save(filename)

    row, col = find_marker_cell(sheet, 'Identifiant', row)
    group_types = GroupType.objects.filter(department=department)
    # if we have too many group types, avoid to destroy the marker below!
    insert_missing_rows(sheet, row, group_types, 5)
    for gt in group_types:
        row = row + 1
        sheet.cell(row=row, column=col, value=gt.name)

    row, col = find_marker_cell(sheet, 'Identifiant', row)
    structural_groups = StructuralGroup.objects.filter(train_prog__in = train_progs)
    insert_missing_rows(sheet, row, structural_groups, 18)
    for gp in structural_groups:
        row = row + 1
        sheet.cell(row=row, column=col, value=gp.name)
        sheet.cell(row=row, column=col+1, value=gp.train_prog.abbrev)
        sheet.cell(row=row, column=col+2, value=gp.type.name)
        if gp.parent_groups.count() == 1:
            sheet.cell(row=row, column=col+3, value=gp.parent_groups.first().name)

    row, col = find_marker_cell(sheet, 'Identifiant', row)
    transversal_groups = TransversalGroup.objects.filter(train_prog__in = train_progs)
    insert_missing_rows(sheet, row, transversal_groups, 18)
    for gp in transversal_groups:
        conflicting_groups = gp.conflicting_groups
        parallel_groups = gp.parallel_groups
        nb_conflicts = conflicting_groups.count()
        if nb_conflicts > 7:
            raise Exception("Trop de groupes en conflit avec %s" % gp)
        row = row + 1
        sheet.cell(row=row, column=col, value=gp.name)
        sheet.cell(row=row, column=col+1, value=gp.train_prog.abbrev)
        col += 2
        for i, confl_group in enumerate(conflicting_groups.all()):
            sheet.cell(row=row, column=col + 2 + i, value=confl_group.name)
        for i, parallel_group in enumerate(parallel_groups.all()):
            sheet.cell(row=row, column=col + 9 + i, value=parallel_group.name)


    sheet = wb[modules_sheet]
    row, col = find_marker_cell(sheet, 'Identifiant')
    modules = Module.objects.filter(train_prog__in=train_progs)
    insert_missing_rows(sheet, row, modules, 27)
    for module in modules:
        row = row + 1
        sheet.cell(row=row, column=col, value=module.abbrev)
        sheet.cell(row=row, column=col+1, value=module.abbrev)
        sheet.cell(row=row, column=col+2, value=module.ppn)
        sheet.cell(row=row, column=col+3, value=module.name)
        sheet.cell(row=row, column=col+4, value=module.train_prog.abbrev)
        sheet.cell(row=row, column=col+5, value=module.period.name)
        if module.head is not None:
            sheet.cell(row=row, column=col+6, value=module.head.username)


    sheet = wb[courses_sheet]
    row, col_start = find_marker_cell(sheet, 'Type')
    course_types = CourseType.objects.filter(department=department)
    insert_missing_rows(sheet, row, list(course_types)+list(course_types), 12)
    for course_type in course_types:
        row = row + 1
        col = col_start
        sheet.cell(row=row, column=col, value=course_type.name)
        sheet.cell(row=row, column=col+1, value=course_type.duration)
        col = col_start + 1
        for group_type in course_type.group_types.all():
            col = col + 1
            sheet.cell(row=row, column=col, value=group_type.name)
        row = row + 1
        col = col_start + 1
        allowed_start_times = course_type.coursestarttimeconstraint_set.first().allowed_start_times
        allowed_start_times.sort()
        for start_time in allowed_start_times:
            col = col + 1
            sheet.cell(row=row, column=col, value=time_from_integer(start_time))

    wb.save(filename)
